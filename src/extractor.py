from typing import List, Dict, Any, Tuple
import io
import re

from openpyxl import load_workbook

# --- Regex auxiliares ---
CURRENCY_RX = re.compile(r"R\$\s*")
SPACES_RX = re.compile(r"\s+")
NON_DIGIT_RX = re.compile(r"[^\d]")  # usado no parse_int_flex

# Cabeçalhos canônicos:
# - REQUIRED_KEYS: obrigatórios na planilha (se faltarem, erro)
# - OPTIONAL_KEYS: opcionais; se existir, usamos, senão calculamos
REQUIRED_KEYS = ["ITEM", "DESCRIÇÃO", "UNID.", "QUANT.", "VALOR UNIT."]
OPTIONAL_KEYS = ["VALOR TOTAL"]
ALL_KEYS = REQUIRED_KEYS + OPTIONAL_KEYS

HEADER_ALIASES = {
    "ITEM": {"ITEM", "ITEN", "ITENS"},
    "DESCRIÇÃO": {
        "DESCRIÇÃO",
        "DESCRICAO",
        "DESCRIÇÃO/ESPECIFICAÇÃO",
        "DESCRIÇÃO DO ITEM",
    },
    "UNID.": {"UNID.", "UNIDADE", "UND", "UNID"},
    "QUANT.": {"QUANT.", "QTD", "QUANTIDADE"},
    "VALOR UNIT.": {
        "VALOR UNIT.",
        "VALOR UNITARIO",
        "VALOR UNITÁRIO",
        "VLR UNIT.",
        "PREÇO UNIT.",
        "PREÇO UNIT",
        "V. UNIT.",   # usado na planilha
        "V UNIT.",
    },
    "VALOR TOTAL": {
        "VALOR TOTAL",
        "VLR TOTAL",
        "TOTAL",
        "PREÇO TOTAL",
        "V. TOTAL",
        "V TOTAL",
    },
}

INCH_VARIANTS = ['"', '\\"', "”", "“", "″"]
INCH_GLYPH = "″"

VALOR_TOTAL_SUFFIX_RX = re.compile(
    r"(VALOR\s+TOTAL\s*R\$\s*[\d\.\,]+)$", flags=re.IGNORECASE
)
VALOR_PRECO_TRAILING_RX = re.compile(r"(R\$\s*[\d\.\,]+)$", flags=re.IGNORECASE)


# -------------------- Utilitários --------------------
def normalize_inches(text: str) -> str:
    if not text:
        return text
    for v in INCH_VARIANTS:
        text = text.replace(v, '"')
    text = re.sub(r"\s+\"", "\"", text)
    text = re.sub(
        r"(\d+(?:[\.,]\d+)?(?:/\d+(?:[\.,]\d+)?)?)\"",
        r"\1" + INCH_GLYPH,
        text,
    )
    return text.replace('"', INCH_GLYPH)


def normalize_text(value) -> str:
    """Normaliza qualquer coisa para string de forma segura (para texto/cabeçalho)."""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    value = value.replace("\xa0", " ")
    value = SPACES_RX.sub(" ", value).strip()
    return normalize_inches(value)


def strip_valor_total_from_desc(text: str) -> str:
    if not text:
        return text
    text = VALOR_TOTAL_SUFFIX_RX.sub("", text).strip()
    text = VALOR_PRECO_TRAILING_RX.sub("", text).strip()
    return text


def parse_money(value):
    """
    Converte valores monetários em float, aceitando:
    - 2.277,92
    - 2277,92
    - 49.52
    - números já como int/float vindos do Excel
    """
    if value is None:
        return None

    # Se já vier numérico do Excel, só converte
    if isinstance(value, (int, float)):
        return float(value)

    s = normalize_text(value)
    if not s:
        return None

    # remove "R$" e caracteres estranhos
    s = CURRENCY_RX.sub("", s).strip()
    allowed = set("0123456789-.,")
    s = "".join(ch for ch in s if ch in allowed)

    if not s or s in {".", ",", "-", "-.", "-,"}:
        return None

    if "," in s and "." in s:
        # formato BR: . milhar, , decimal
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif "," in s:
        # só vírgula: assume vírgula decimal
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        # só ponto ou nenhum separador: já é formato padrão
        pass

    try:
        return float(s)
    except ValueError:
        return None


def parse_int_flex(value):
    """
    Converte ITEM/QUANT em int:
    - aceita int/float direto do Excel
    - aceita strings tipo "180", "180.0", " 426 ", etc.
    """
    if value is None:
        return None

    # numérico direto
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))

    s = normalize_text(value)
    if not s:
        return None

    # tenta via float primeiro (para "180.0" / "426,0")
    try:
        f = float(s.replace(",", "."))
        return int(round(f))
    except ValueError:
        pass

    # fallback: mantém só dígitos
    only_digits = NON_DIGIT_RX.sub("", s)
    if not only_digits:
        return None
    try:
        return int(only_digits)
    except ValueError:
        return None


def looks_like_total_row(row: List[Any]) -> bool:
    joined = " ".join(normalize_text(c).upper() for c in row)
    return "VALOR TOTAL" in joined and parse_money(joined) is not None


# -------------------- Mapeamento --------------------
def _header_match_score(cell, target_key: str) -> int:
    cell_u = normalize_text(cell).upper()
    return int(any(alias in cell_u for alias in HEADER_ALIASES[target_key]))


def _find_col_index(header_row: List[Any], key: str):
    for j, cell in enumerate(header_row):
        cell_u = normalize_text(cell).upper()
        for alias in HEADER_ALIASES[key]:
            if alias in cell_u:
                return j
    return None


def detect_header_and_map(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    """
    Retorna (índice_da_linha_de_cabeçalho, col_map).
    Exige somente as colunas de REQUIRED_KEYS.
    """
    best_idx = -1
    best_score = -1

    for i, row in enumerate(rows):
        score = sum(
            any(_header_match_score(c, key) for c in row) for key in ALL_KEYS
        )
        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx < 0 or best_score < 3:  # 3+ chaves já é bem forte
        raise ValueError("Cabeçalho da tabela não foi encontrado com confiança suficiente.")

    header_row = rows[best_idx]

    col_map: Dict[str, int] = {}
    # mapeia chaves obrigatórias + opcionais
    for key in ALL_KEYS:
        idx = _find_col_index(header_row, key)
        if idx is not None:
            col_map[key] = idx

    missing_required = [k for k in REQUIRED_KEYS if k not in col_map]
    if missing_required:
        raise ValueError(
            f"Não foi possível mapear as colunas obrigatórias: faltam {missing_required}"
        )

    return best_idx, col_map


# -------------------- Parsing --------------------
def parse_row_with_map(row: List[Any], col_map: Dict[str, int]) -> Dict[str, Any]:
    def get_raw(key: str):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    raw_item = get_raw("ITEM")
    raw_desc = get_raw("DESCRIÇÃO")
    raw_unid = get_raw("UNID.")
    raw_quant = get_raw("QUANT.")
    raw_vu = get_raw("VALOR UNIT.")
    raw_vt = get_raw("VALOR TOTAL")  # pode ser None se coluna não existir

    item = parse_int_flex(raw_item)
    descricao = strip_valor_total_from_desc(normalize_text(raw_desc))
    unid = normalize_text(raw_unid).upper().rstrip(".,")
    quant = parse_int_flex(raw_quant)
    valor_unit = parse_money(raw_vu)
    valor_total = parse_money(raw_vt)

    return {
        "item": item,
        "descricao": descricao,
        "unid": unid,
        "quant": quant,
        "valor_unit": valor_unit,
        "valor_total": valor_total,  # será recalculado em validate_and_fix
    }


def validate_and_fix(rec: Dict[str, Any]):
    vu = rec.get("valor_unit")
    qt = rec.get("quant")

    # Sempre recalcula valor_total a partir de quant * unit
    if vu is not None and qt is not None:
        rec["valor_total"] = round(vu * qt, 2)

    return rec


def build_issues(rec: Dict[str, Any]):
    # valor_total não é mais considerado "faltando", pois é calculado
    missing = [
        k
        for k in ["item", "descricao", "unid", "quant", "valor_unit"]
        if rec.get(k) in ("", None)
    ]
    if missing:
        return {
            "item": rec.get("item"),
            "missing": missing,
        }
    return None


# -------------------- Extração principal --------------------
def _extract_from_all_rows(all_rows: List[List[Any]], empty_msg: str):
    if not all_rows:
        return {"rows": [], "issues": [{"error": empty_msg}]}

    header_idx, col_map = detect_header_and_map(all_rows)

    data_rows: List[List[Any]] = []
    for row in all_rows[header_idx + 1 :]:
        # pula linha vazia
        if not row or not any(normalize_text(c) for c in row):
            continue
        # para na linha de totalização geral (se existir texto VALOR TOTAL em alguma linha)
        if looks_like_total_row(row):
            break
        data_rows.append(row)

    records: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []

    for row in data_rows:
        rec = parse_row_with_map(row, col_map)
        rec = validate_and_fix(rec)

        issue = build_issues(rec)
        if issue:
            issues.append(issue)

        records.append(rec)

    records.sort(key=lambda r: (r["item"] if r["item"] is not None else 10**9))

    return {"rows": records, "issues": issues}


def extract_table_from_xlsx(file_bytes: bytes):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return {
            "rows": [],
            "issues": [{"error": "Não foi possível ler o arquivo XLSX."}],
        }

    sheet = wb.active

    all_rows: List[List[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        raw_row = list(row)
        # normaliza para ver se é vazia, mas guarda tipos originais
        norm_row = [normalize_text(c) for c in raw_row]
        if any(norm_row):
            all_rows.append(raw_row)

    return _extract_from_all_rows(all_rows, "Nenhuma tabela encontrada na planilha XLSX.")
